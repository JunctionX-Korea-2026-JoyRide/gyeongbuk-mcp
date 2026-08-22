"""Build a compact SQLite snapshot from downloaded public-data files."""

from __future__ import annotations

import csv
import hashlib
import math
import re
import sqlite3
import tempfile
import unicodedata
from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class BuildSummary:
    """Counts and output location for one successful snapshot build."""

    database_path: Path
    hospitals: int
    departments: int
    bus_stops: int
    bus_routes: int
    bus_stop_routes: int
    bus_frequencies: int
    bus_route_patterns: int
    bus_pattern_stops: int
    bus_pattern_frequencies: int
    bus_frequency_routes: int
    markets: int
    population_age_bands: int
    safety_grades: int


def build_snapshot(raw_dir: Path, reference_dir: Path, output_path: Path) -> BuildSummary:
    """Normalize the required raw files and atomically replace the SQLite snapshot."""

    raw_dir = raw_dir.resolve()
    reference_dir = reference_dir.resolve()
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    hira_zip = _find_hira_zip(raw_dir)
    bus_stop_csv = _find_csv(raw_dir, {"정류장번호", "정류장명", "도시코드"})
    pohang_route_csv = _find_csv(raw_dir, {"노선명", "승강장명칭", "운행시간"})
    market_csv = _find_csv(raw_dir, {"시장명", "시장유형", "소재지도로명주소"})
    population_csv = _find_population_csv(raw_dir)
    safety_hwpx = _find_safety_hwpx(raw_dir)
    frequency_csv = reference_dir / "pohang_bus_frequencies.csv"
    pattern_frequency_csv = reference_dir / "pohang_branch_pattern_frequencies.csv"
    if not frequency_csv.is_file():
        raise FileNotFoundError(f"버스 배차 기준 파일이 없습니다: {frequency_csv}")
    if not pattern_frequency_csv.is_file():
        raise FileNotFoundError(f"지선 운행횟수 검수 파일이 없습니다: {pattern_frequency_csv}")

    with tempfile.NamedTemporaryFile(
        prefix="gyeongbuk-", suffix=".sqlite3", dir=output_path.parent, delete=False
    ) as temporary:
        temporary_path = Path(temporary.name)
    try:
        with sqlite3.connect(temporary_path) as connection:
            _create_schema(connection)
            hospitals, departments = _load_hospitals(connection, hira_zip)
            bus_stops = _load_bus_stops(connection, bus_stop_csv)
            (
                bus_routes,
                bus_stop_routes,
                bus_route_patterns,
                bus_pattern_stops,
            ) = _load_pohang_routes(connection, pohang_route_csv)
            bus_frequencies = _load_bus_frequencies(connection, frequency_csv)
            bus_pattern_frequencies = _load_bus_pattern_frequencies(
                connection, pattern_frequency_csv, raw_dir
            )
            bus_frequency_routes = _validate_bus_frequency_coverage(connection)
            markets = _load_markets(connection, market_csv)
            population_age_bands = _load_population_age_bands(connection, population_csv)
            safety_grades = _load_safety_grades(connection, safety_hwpx)
            _write_metadata(
                connection,
                [
                    hira_zip,
                    bus_stop_csv,
                    pohang_route_csv,
                    market_csv,
                    frequency_csv,
                    pattern_frequency_csv,
                    population_csv,
                    safety_hwpx,
                    *sorted(
                        raw_dir / name for name in _referenced_pdf_names(pattern_frequency_csv)
                    ),
                ],
            )
            connection.commit()
        temporary_path.replace(output_path)
    finally:
        temporary_path.unlink(missing_ok=True)

    return BuildSummary(
        database_path=output_path,
        hospitals=hospitals,
        departments=departments,
        bus_stops=bus_stops,
        bus_routes=bus_routes,
        bus_stop_routes=bus_stop_routes,
        bus_frequencies=bus_frequencies,
        bus_route_patterns=bus_route_patterns,
        bus_pattern_stops=bus_pattern_stops,
        bus_pattern_frequencies=bus_pattern_frequencies,
        bus_frequency_routes=bus_frequency_routes,
        markets=markets,
        population_age_bands=population_age_bands,
        safety_grades=safety_grades,
    )


def _create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        PRAGMA journal_mode = DELETE;
        PRAGMA foreign_keys = ON;

        CREATE TABLE metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE hospitals (
            institution_id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            address TEXT NOT NULL,
            phone TEXT,
            institution_type TEXT,
            latitude REAL NOT NULL,
            longitude REAL NOT NULL
        );
        CREATE TABLE hospital_departments (
            institution_id TEXT NOT NULL REFERENCES hospitals(institution_id),
            department_code TEXT NOT NULL,
            department_name TEXT NOT NULL,
            PRIMARY KEY (institution_id, department_code)
        );
        CREATE TABLE bus_stops (
            stop_id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            city_code TEXT NOT NULL,
            city_name TEXT NOT NULL,
            latitude REAL NOT NULL,
            longitude REAL NOT NULL,
            reference_date TEXT
        );
        CREATE TABLE bus_routes (
            route_id TEXT PRIMARY KEY,
            route_number TEXT NOT NULL,
            route_type TEXT,
            city_code TEXT NOT NULL,
            reference_date TEXT
        );
        CREATE TABLE bus_stop_routes (
            stop_id TEXT NOT NULL REFERENCES bus_stops(stop_id),
            route_id TEXT NOT NULL REFERENCES bus_routes(route_id),
            PRIMARY KEY (stop_id, route_id)
        );
        CREATE TABLE bus_route_frequencies (
            route_id TEXT PRIMARY KEY REFERENCES bus_routes(route_id),
            first_bus TEXT,
            last_bus TEXT,
            interval_weekday INTEGER,
            interval_saturday INTEGER,
            interval_sunday INTEGER,
            trips_weekday INTEGER,
            trips_saturday INTEGER,
            trips_sunday INTEGER,
            frequency_basis TEXT NOT NULL CHECK (
                frequency_basis IN (
                    'published_trip_count',
                    'conservative_interval_estimate',
                    'api_summary'
                )
            ),
            reference_date TEXT,
            source_url TEXT NOT NULL
        );
        CREATE TABLE bus_route_patterns (
            pattern_id TEXT PRIMARY KEY,
            route_id TEXT NOT NULL REFERENCES bus_routes(route_id),
            route_detail TEXT NOT NULL,
            UNIQUE (route_id, route_detail)
        );
        CREATE TABLE bus_pattern_stops (
            pattern_id TEXT NOT NULL REFERENCES bus_route_patterns(pattern_id),
            stop_id TEXT NOT NULL REFERENCES bus_stops(stop_id),
            stop_sequence INTEGER NOT NULL CHECK (stop_sequence >= 0),
            PRIMARY KEY (pattern_id, stop_id, stop_sequence)
        );
        CREATE TABLE bus_pattern_frequencies (
            schedule_id TEXT PRIMARY KEY,
            pattern_id TEXT NOT NULL REFERENCES bus_route_patterns(pattern_id),
            first_bus TEXT NOT NULL,
            last_bus TEXT NOT NULL,
            trips_weekday INTEGER NOT NULL CHECK (trips_weekday >= 0),
            trips_saturday INTEGER NOT NULL CHECK (trips_saturday >= 0),
            trips_sunday INTEGER NOT NULL CHECK (trips_sunday >= 0),
            frequency_basis TEXT NOT NULL CHECK (frequency_basis = 'published_trip_count'),
            reference_date TEXT NOT NULL,
            source_document TEXT NOT NULL,
            source_page INTEGER NOT NULL CHECK (source_page >= 1),
            source_url TEXT NOT NULL,
            notes TEXT
        );
        CREATE TABLE markets (
            market_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            market_type TEXT,
            road_address TEXT,
            lot_address TEXT,
            opening_cycle TEXT,
            latitude REAL NOT NULL,
            longitude REAL NOT NULL,
            reference_date TEXT,
            UNIQUE (name, road_address, lot_address)
        );
        CREATE TABLE population_age_bands (
            region_code TEXT NOT NULL,
            region_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            region_level TEXT NOT NULL,
            as_of TEXT NOT NULL,
            age_from INTEGER NOT NULL,
            age_to INTEGER NOT NULL,
            population INTEGER NOT NULL,
            total_population INTEGER NOT NULL,
            PRIMARY KEY (region_code, as_of, age_from, age_to)
        );
        CREATE TABLE safety_grades (
            region_name TEXT NOT NULL,
            normalized_name TEXT NOT NULL,
            region_level TEXT NOT NULL,
            category TEXT NOT NULL,
            grade INTEGER NOT NULL CHECK (grade BETWEEN 1 AND 5),
            publication_year INTEGER NOT NULL,
            statistics_year INTEGER NOT NULL,
            comparison_group TEXT NOT NULL,
            PRIMARY KEY (normalized_name, category, publication_year)
        );

        CREATE INDEX hospitals_location ON hospitals(latitude, longitude);
        CREATE INDEX departments_code ON hospital_departments(department_code);
        CREATE INDEX bus_stops_location ON bus_stops(latitude, longitude);
        CREATE INDEX bus_stops_name ON bus_stops(normalized_name);
        CREATE INDEX bus_patterns_route ON bus_route_patterns(route_id);
        CREATE INDEX bus_pattern_stops_stop ON bus_pattern_stops(stop_id, pattern_id);
        CREATE INDEX bus_pattern_frequencies_pattern ON bus_pattern_frequencies(pattern_id);
        CREATE INDEX markets_location ON markets(latitude, longitude);
        CREATE INDEX markets_addresses ON markets(road_address, lot_address);
        CREATE INDEX population_region_name ON population_age_bands(normalized_name, as_of);
        CREATE INDEX safety_region_name ON safety_grades(normalized_name, publication_year);
        """
    )


def _load_hospitals(connection: sqlite3.Connection, archive_path: Path) -> tuple[int, int]:
    with ZipFile(archive_path) as archive:
        hospital_name = next(
            name for name in archive.namelist() if name.rsplit("/", 1)[-1].startswith("1.")
        )
        department_name = next(
            name for name in archive.namelist() if name.rsplit("/", 1)[-1].startswith("5.")
        )

        hospital_rows: list[tuple[object, ...]] = []
        workbook = load_workbook(
            BytesIO(archive.read(hospital_name)), read_only=True, data_only=True
        )
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError("병원정보 XLSX에 워크시트가 없습니다.")
        iterator = worksheet.iter_rows(values_only=True)
        header = _header_map(next(iterator))
        for row in iterator:
            if _cell(row, header, "시도코드명") != "경북":
                continue
            latitude = _float(_cell(row, header, "좌표(Y)"))
            longitude = _float(_cell(row, header, "좌표(X)"))
            if latitude is None or longitude is None:
                continue
            hospital_rows.append(
                (
                    _text(_cell(row, header, "암호화요양기호")),
                    _text(_cell(row, header, "요양기관명")),
                    _text(_cell(row, header, "주소")),
                    _optional_text(_cell(row, header, "전화번호")),
                    _optional_text(_cell(row, header, "종별코드명")),
                    latitude,
                    longitude,
                )
            )
        workbook.close()
        connection.executemany("INSERT INTO hospitals VALUES (?, ?, ?, ?, ?, ?, ?)", hospital_rows)

        institution_ids = {str(row[0]) for row in hospital_rows}
        department_rows: list[tuple[str, str, str]] = []
        workbook = load_workbook(
            BytesIO(archive.read(department_name)), read_only=True, data_only=True
        )
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError("진료과목 XLSX에 워크시트가 없습니다.")
        iterator = worksheet.iter_rows(values_only=True)
        header = _header_map(next(iterator))
        seen: set[tuple[str, str]] = set()
        for row in iterator:
            institution_id = _text(_cell(row, header, "암호화요양기호"))
            department_code = _text(_cell(row, header, "진료과목코드"))
            key = (institution_id, department_code)
            if institution_id not in institution_ids or key in seen:
                continue
            seen.add(key)
            department_rows.append(
                (institution_id, department_code, _text(_cell(row, header, "진료과목코드명")))
            )
        workbook.close()
        connection.executemany("INSERT INTO hospital_departments VALUES (?, ?, ?)", department_rows)
    return len(hospital_rows), len(department_rows)


def _load_bus_stops(connection: sqlite3.Connection, path: Path) -> int:
    rows: list[tuple[object, ...]] = []
    with path.open(encoding="cp949", newline="") as stream:
        for row in csv.DictReader(stream):
            if not row["도시명"].startswith("경상북도"):
                continue
            latitude = _float(row["위도"])
            longitude = _float(row["경도"])
            if latitude is None or longitude is None:
                continue
            rows.append(
                (
                    row["정류장번호"].strip(),
                    row["정류장명"].strip(),
                    _normalize_name(row["정류장명"]),
                    row["도시코드"].strip(),
                    row["도시명"].strip(),
                    latitude,
                    longitude,
                    row["정보수집일"].strip() or None,
                )
            )
    connection.executemany("INSERT OR IGNORE INTO bus_stops VALUES (?, ?, ?, ?, ?, ?, ?, ?)", rows)
    return int(connection.execute("SELECT COUNT(*) FROM bus_stops").fetchone()[0])


def _load_pohang_routes(connection: sqlite3.Connection, path: Path) -> tuple[int, int, int, int]:
    with path.open(encoding="cp949", newline="") as stream:
        raw_rows = list(csv.DictReader(stream))
    route_numbers = sorted({row["노선명"].strip() for row in raw_rows if row["노선명"].strip()})
    reference_date = max((row["데이터기준일자"].strip() for row in raw_rows), default=None)
    route_rows = [
        (f"pohang:{number}", number, "포항시 시내버스", "37010", reference_date)
        for number in route_numbers
    ]
    connection.executemany("INSERT INTO bus_routes VALUES (?, ?, ?, ?, ?)", route_rows)

    route_details = sorted(
        {
            (row["노선명"].strip(), row["노선상세"].strip())
            for row in raw_rows
            if row["노선명"].strip() and row["노선상세"].strip()
        }
    )
    pattern_ids = {key: _pattern_id(*key) for key in route_details}
    connection.executemany(
        "INSERT INTO bus_route_patterns VALUES (?, ?, ?)",
        [
            (pattern_ids[(route_number, route_detail)], f"pohang:{route_number}", route_detail)
            for route_number, route_detail in route_details
        ],
    )

    stop_name_rows = connection.execute(
        "SELECT stop_id, normalized_name FROM bus_stops WHERE city_code = '37010'"
    ).fetchall()
    stop_ids_by_name: dict[str, list[str]] = {}
    for stop_id, normalized_name in stop_name_rows:
        stop_ids_by_name.setdefault(str(normalized_name), []).append(str(stop_id))

    associations: set[tuple[str, str]] = set()
    pattern_stops: set[tuple[str, str, int]] = set()
    for row in raw_rows:
        route_number = row["노선명"].strip()
        route_detail = row["노선상세"].strip()
        normalized_stop = _normalize_name(row["승강장명칭"])
        stop_sequence = _int(row["승강장순번"])
        if stop_sequence is None or stop_sequence < 0:
            raise ValueError(
                f"포항 노선 CSV의 승강장순번이 올바르지 않습니다: "
                f"{route_number}/{route_detail}/{row['승강장순번']}"
            )
        for stop_id in stop_ids_by_name.get(normalized_stop, []):
            associations.add((stop_id, f"pohang:{route_number}"))
            pattern_stops.add((pattern_ids[(route_number, route_detail)], stop_id, stop_sequence))
    connection.executemany("INSERT INTO bus_stop_routes VALUES (?, ?)", sorted(associations))
    connection.executemany("INSERT INTO bus_pattern_stops VALUES (?, ?, ?)", sorted(pattern_stops))
    return len(route_rows), len(associations), len(route_details), len(pattern_stops)


def _load_bus_frequencies(connection: sqlite3.Connection, path: Path) -> int:
    rows: list[tuple[object, ...]] = []
    with path.open(encoding="utf-8", newline="") as stream:
        for row in csv.DictReader(stream):
            maximum_interval = _int(row["maximum_interval_minutes"])
            explicit_trips = _int(row["explicit_daily_trips"])
            trips = (
                explicit_trips
                if explicit_trips is not None
                else _conservative_trips(row["first_bus"], row["last_bus"], maximum_interval)
            )
            if trips is None:
                raise ValueError(
                    f"간선 배차 기준에 운행횟수를 계산할 정보가 없습니다: {row['route_number']}"
                )
            rows.append(
                (
                    f"pohang:{row['route_number']}",
                    row["first_bus"] or None,
                    row["last_bus"] or None,
                    maximum_interval,
                    maximum_interval,
                    maximum_interval,
                    trips,
                    trips,
                    trips,
                    (
                        "published_trip_count"
                        if explicit_trips is not None
                        else "conservative_interval_estimate"
                    ),
                    row["as_of"],
                    row["source_url"],
                )
            )
    connection.executemany(
        "INSERT INTO bus_route_frequencies VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", rows
    )
    return len(rows)


def _load_bus_pattern_frequencies(connection: sqlite3.Connection, path: Path, raw_dir: Path) -> int:
    required_columns = {
        "schedule_id",
        "route_number",
        "route_detail",
        "first_bus",
        "last_bus",
        "trips_weekday",
        "trips_saturday",
        "trips_sunday",
        "reference_date",
        "source_document",
        "source_page",
        "source_url",
        "notes",
    }
    with path.open(encoding="utf-8", newline="") as stream:
        reader = csv.DictReader(stream)
        if reader.fieldnames is None or not required_columns <= set(reader.fieldnames):
            missing = sorted(required_columns - set(reader.fieldnames or []))
            raise ValueError(f"지선 운행횟수 검수 CSV 필수 열이 없습니다: {', '.join(missing)}")
        source_rows = list(reader)

    known_patterns = {
        (str(route_number), str(route_detail)): (str(pattern_id), str(route_id))
        for pattern_id, route_id, route_number, route_detail in connection.execute(
            "SELECT p.pattern_id, p.route_id, r.route_number, p.route_detail "
            "FROM bus_route_patterns p JOIN bus_routes r ON r.route_id = p.route_id"
        )
    }
    known_route_frequency_ids = {
        str(row[0]) for row in connection.execute("SELECT route_id FROM bus_route_frequencies")
    }
    seen_schedule_ids: set[str] = set()
    rows: list[tuple[object, ...]] = []
    for line_number, row in enumerate(source_rows, start=2):
        schedule_id = row["schedule_id"].strip()
        route_number = row["route_number"].strip()
        route_detail = row["route_detail"].strip()
        if not schedule_id:
            raise ValueError(f"지선 운행횟수 검수 CSV {line_number}행의 schedule_id가 비었습니다.")
        if schedule_id in seen_schedule_ids:
            raise ValueError(f"중복 schedule_id입니다: {schedule_id}")
        seen_schedule_ids.add(schedule_id)
        pattern = known_patterns.get((route_number, route_detail))
        if pattern is None:
            raise ValueError(
                f"포항 노선 CSV에 없는 노선·노선상세입니다: {route_number}/{route_detail}"
            )
        pattern_id, route_id = pattern
        if route_id in known_route_frequency_ids:
            raise ValueError(f"간선 배차와 지선 패턴 배차가 중복됩니다: {route_number}")

        source_document = row["source_document"].strip()
        if Path(source_document).name != source_document or not source_document.lower().endswith(
            ".pdf"
        ):
            raise ValueError(
                f"source_document는 data/raw의 PDF 파일명이어야 합니다: {source_document}"
            )
        source_path = raw_dir / source_document
        if not source_path.is_file() or source_path.read_bytes()[:5] != b"%PDF-":
            raise FileNotFoundError(f"검수 CSV가 참조하는 원본 PDF가 없습니다: {source_path}")

        first_bus = row["first_bus"].strip()
        last_bus = row["last_bus"].strip()
        if _clock_minutes(first_bus) is None or _clock_minutes(last_bus) is None:
            raise ValueError(f"잘못된 첫차·막차 형식입니다: {schedule_id}")
        trip_counts = tuple(
            _strict_nonnegative_int(row[name], schedule_id, name)
            for name in (
                "trips_weekday",
                "trips_saturday",
                "trips_sunday",
            )
        )
        source_page = _strict_nonnegative_int(row["source_page"], schedule_id, "source_page")
        if source_page < 1:
            raise ValueError(f"source_page는 1 이상이어야 합니다: {schedule_id}")
        reference_date = row["reference_date"].strip()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", reference_date) is None:
            raise ValueError(f"reference_date는 YYYY-MM-DD 형식이어야 합니다: {schedule_id}")
        source_url = row["source_url"].strip()
        if not source_url.startswith(("https://", "http://")):
            raise ValueError(f"source_url이 올바르지 않습니다: {schedule_id}")
        rows.append(
            (
                schedule_id,
                pattern_id,
                first_bus,
                last_bus,
                *trip_counts,
                "published_trip_count",
                reference_date,
                source_document,
                source_page,
                source_url,
                row["notes"].strip() or None,
            )
        )
    connection.executemany(
        "INSERT INTO bus_pattern_frequencies VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        rows,
    )
    return len(rows)


def _validate_bus_frequency_coverage(connection: sqlite3.Connection) -> int:
    covered = {
        str(row[0])
        for row in connection.execute(
            "SELECT r.route_number FROM bus_routes r JOIN bus_route_frequencies f "
            "ON f.route_id = r.route_id UNION SELECT r.route_number FROM bus_routes r "
            "JOIN bus_route_patterns p ON p.route_id = r.route_id "
            "JOIN bus_pattern_frequencies f ON f.pattern_id = p.pattern_id"
        )
    }
    all_routes = {str(row[0]) for row in connection.execute("SELECT route_number FROM bus_routes")}
    missing = all_routes - covered
    expected_missing = {"임시노선", "죽장DRT"}
    if len(all_routes) != 53 or len(covered) != 51 or missing != expected_missing:
        raise ValueError(
            "포항 배차 판정 범위가 기대값과 다릅니다: "
            f"전체={len(all_routes)}, 판정={len(covered)}, 미상={sorted(missing)}"
        )
    return len(covered)


def _referenced_pdf_names(path: Path) -> set[str]:
    with path.open(encoding="utf-8", newline="") as stream:
        return {
            row["source_document"].strip()
            for row in csv.DictReader(stream)
            if row.get("source_document", "").strip()
        }


def _load_markets(connection: sqlite3.Connection, path: Path) -> int:
    rows: list[tuple[object, ...]] = []
    with path.open(encoding=_csv_encoding(path), newline="") as stream:
        for row in csv.DictReader(stream):
            combined_address = f"{row['소재지도로명주소']} {row['소재지지번주소']}"
            if "경상북도" not in combined_address:
                continue
            latitude = _float(row["위도"])
            longitude = _float(row["경도"])
            if latitude is None or longitude is None:
                continue
            rows.append(
                (
                    row["시장명"].strip(),
                    row["시장유형"].strip() or None,
                    row["소재지도로명주소"].strip() or None,
                    row["소재지지번주소"].strip() or None,
                    row["시장개설주기"].strip() or None,
                    latitude,
                    longitude,
                    row["데이터기준일자"].strip() or None,
                )
            )
    connection.executemany(
        "INSERT OR IGNORE INTO markets "
        "(name, market_type, road_address, lot_address, opening_cycle, latitude, longitude, "
        "reference_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        rows,
    )
    return int(connection.execute("SELECT COUNT(*) FROM markets").fetchone()[0])


def _load_population_age_bands(connection: sqlite3.Connection, path: Path) -> int:
    rows: list[tuple[object, ...]] = []
    with path.open(encoding="cp949", newline="") as stream:
        reader = csv.DictReader(stream)
        if reader.fieldnames is None:
            raise ValueError("연령별 인구 CSV에 헤더가 없습니다.")
        total_column = next(
            (name for name in reader.fieldnames if name.endswith("_계_총인구수")), None
        )
        if total_column is None:
            raise ValueError("연령별 인구 CSV에서 총인구수 열을 찾지 못했습니다.")
        as_of_match = re.match(r"(\d{4})년(\d{2})월", total_column)
        if as_of_match is None:
            raise ValueError("연령별 인구 CSV의 기준년월을 해석하지 못했습니다.")
        as_of = "".join(as_of_match.groups())
        age_columns: list[tuple[str, int, int]] = []
        for name in reader.fieldnames:
            match = re.search(r"_계_(\d+)~(\d+)세$", name)
            if match:
                age_columns.append((name, int(match.group(1)), int(match.group(2))))
                continue
            match = re.search(r"_계_(\d+)세 이상$", name)
            if match:
                age_columns.append((name, int(match.group(1)), 130))

        for row in reader:
            region_match = re.match(r"\s*(.*?)\s*\((\d{10})\)\s*$", row["행정구역"])
            if region_match is None:
                continue
            region_name, region_code = region_match.groups()
            if not region_name.startswith("경상북도"):
                continue
            total_population = _int(row[total_column])
            if total_population is None:
                continue
            region_level = _population_region_level(region_name)
            for column, age_from, age_to in age_columns:
                population = _int(row[column])
                if population is None:
                    continue
                rows.append(
                    (
                        region_code,
                        region_name,
                        _normalize_name(region_name),
                        region_level,
                        as_of,
                        age_from,
                        age_to,
                        population,
                        total_population,
                    )
                )
    connection.executemany(
        "INSERT INTO population_age_bands VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", rows
    )
    return len(rows)


def _load_safety_grades(connection: sqlite3.Connection, path: Path) -> int:
    category_names = {
        "교통사고": "traffic_accident",
        "화재": "fire",
        "범죄": "crime",
        "생활안전": "life_safety",
        "자살": "suicide",
        "감염병": "infectious_disease",
    }
    rows: list[tuple[object, ...]] = []
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("Contents/section0.xml"))
    for table in (element for element in root.iter() if element.tag.endswith("}tbl")):
        table_rows = _hwpx_table_rows(table)
        if not table_rows:
            continue
        header = table_rows[0]
        if "범죄" not in header or "시도" not in header:
            continue
        header_index = {name: index for index, name in enumerate(header)}
        for values in table_rows[1:]:
            if len(values) != len(header):
                continue
            province = values[header_index["시도"]]
            district_index = header_index.get("시군구")
            if district_index is None:
                if province != "경상북도":
                    continue
                region_name = province
                region_level = "province"
                comparison_group = "도"
            else:
                if province != "경북":
                    continue
                district = values[district_index]
                region_name = f"경상북도 {district}"
                region_level = "city_county"
                comparison_group = _safety_comparison_group(district)
            for korean_name, category in category_names.items():
                grade = _int(values[header_index[korean_name]])
                if grade is None or not 1 <= grade <= 5:
                    continue
                rows.append(
                    (
                        region_name,
                        _normalize_name(region_name),
                        region_level,
                        category,
                        grade,
                        2025,
                        2024,
                        comparison_group,
                    )
                )
    connection.executemany("INSERT INTO safety_grades VALUES (?, ?, ?, ?, ?, ?, ?, ?)", rows)
    return len(rows)


def _hwpx_table_rows(table: ElementTree.Element) -> list[list[str]]:
    rows: list[list[str]] = []
    for table_row in (element for element in table.iter() if element.tag.endswith("}tr")):
        cells: list[str] = []
        for cell in (element for element in table_row if element.tag.endswith("}tc")):
            text = "".join(
                element.text or "" for element in cell.iter() if element.tag.endswith("}t")
            ).strip()
            cells.append(text)
        if cells:
            rows.append(cells)
    return rows


def _population_region_level(region_name: str) -> str:
    parts = region_name.split()
    if len(parts) == 1:
        return "province"
    if len(parts) == 2:
        return "city_county"
    return "town"


def _safety_comparison_group(district: str) -> str:
    if district.endswith("시"):
        return "시"
    if district.endswith("군"):
        return "군"
    return "구"


def _write_metadata(connection: sqlite3.Connection, paths: list[Path]) -> None:
    metadata = {
        "built_at": datetime.now(UTC).isoformat(),
        "schema_version": "3",
    }
    for path in paths:
        metadata[f"sha256:{path.name}"] = _sha256(path)
    connection.executemany("INSERT INTO metadata VALUES (?, ?)", sorted(metadata.items()))


def _find_hira_zip(raw_dir: Path) -> Path:
    for path in sorted(raw_dir.glob("*.zip")):
        with ZipFile(path) as archive:
            names = archive.namelist()
            if any(name.rsplit("/", 1)[-1].startswith("1.") for name in names) and any(
                name.rsplit("/", 1)[-1].startswith("5.") for name in names
            ):
                return path
    raise FileNotFoundError("병원정보와 진료과목 XLSX를 포함한 HIRA ZIP을 찾지 못했습니다.")


def _find_csv(raw_dir: Path, required_columns: set[str]) -> Path:
    for path in sorted(raw_dir.glob("*.csv")):
        try:
            with path.open(encoding=_csv_encoding(path), newline="") as stream:
                header = set(next(csv.reader(stream)))
        except (UnicodeDecodeError, StopIteration):
            continue
        if required_columns <= header:
            return path
    names = ", ".join(sorted(required_columns))
    raise FileNotFoundError(f"필수 열({names})을 가진 CSV를 찾지 못했습니다.")


def _csv_encoding(path: Path) -> str:
    with path.open("rb") as stream:
        return "utf-8-sig" if stream.read(3) == b"\xef\xbb\xbf" else "cp949"


def _find_population_csv(raw_dir: Path) -> Path:
    for path in sorted(raw_dir.glob("*.csv")):
        try:
            with path.open(encoding="cp949", newline="") as stream:
                header = next(csv.reader(stream))
        except (UnicodeDecodeError, StopIteration):
            continue
        if "행정구역" in header and any(name.endswith("_계_70~79세") for name in header):
            return path
    raise FileNotFoundError("행정안전부 연령별 인구 CSV를 찾지 못했습니다.")


def _find_safety_hwpx(raw_dir: Path) -> Path:
    for path in sorted(raw_dir.glob("*.hwpx")):
        try:
            with ZipFile(path) as archive:
                preview = archive.read("Preview/PrvText.txt").decode("utf-8")
        except (KeyError, UnicodeDecodeError):
            continue
        if "지역안전지수 산출 결과" in preview and "범죄" in preview:
            return path
    raise FileNotFoundError("행정안전부 지역안전지수 HWPX를 찾지 못했습니다.")


def _header_map(row: tuple[object, ...]) -> dict[str, int]:
    return {_text(value): index for index, value in enumerate(row)}


def _cell(row: tuple[object, ...], header: dict[str, int], name: str) -> object:
    return row[header[name]]


def _normalize_name(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value)
    return re.sub(r"[^0-9A-Za-z가-힣]", "", normalized).lower()


def _pattern_id(route_number: str, route_detail: str) -> str:
    normalized_detail = _normalize_name(route_detail)
    digest = hashlib.sha256(f"{route_number}\0{route_detail}".encode()).hexdigest()[:16]
    return f"pohang-pattern:{route_number}:{normalized_detail[:32]}:{digest}"


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _optional_text(value: object) -> str | None:
    text = _text(value)
    return text or None


def _float(value: object) -> float | None:
    try:
        return float(_text(value).replace(",", ""))
    except ValueError:
        return None


def _int(value: object) -> int | None:
    number = _float(value)
    return int(number) if number is not None else None


def _strict_nonnegative_int(value: object, schedule_id: str, field_name: str) -> int:
    text = _text(value)
    if re.fullmatch(r"\d+", text) is None:
        raise ValueError(f"{field_name}은 0 이상의 정수여야 합니다: {schedule_id}")
    return int(text)


def _conservative_trips(first: str, last: str, interval: int | None) -> int | None:
    if not first or not last or not interval:
        return None
    first_minutes = _clock_minutes(first)
    last_minutes = _clock_minutes(last)
    if first_minutes is None or last_minutes is None:
        return None
    if last_minutes < first_minutes:
        last_minutes += 24 * 60
    return math.floor((last_minutes - first_minutes) / interval) + 1


def _clock_minutes(value: str) -> int | None:
    digits = re.sub(r"\D", "", value).zfill(4)
    if len(digits) != 4:
        return None
    hour, minute = int(digits[:2]), int(digits[2:])
    return hour * 60 + minute if hour < 24 and minute < 60 else None


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
